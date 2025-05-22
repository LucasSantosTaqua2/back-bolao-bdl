# app/api/v1/endpoints/games.py
from typing import Annotated, List, Any
from datetime import datetime, timezone # Adicione timezone
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session # <<< MUDANÇA: Use Session do SQLAlchemy ORM
from sqlalchemy import select # <<< MUDANÇA: Use select do SQLAlchemy principal
import openpyxl
from io import BytesIO

from app.core.database import get_session
# MUDANÇA: Importe get_current_active_admin e get_current_user do core.security
from app.core.security import get_current_active_admin, get_current_user
from app.models.game import Game, GameStatus
# MUDANÇA: Importe as funções CRUD do seu arquivo app/crud/game.py (que agora está atualizado para SQLAlchemy Puro)
from app.crud.game import (
    create_game,
    get_games_by_round,
    update_game_result,
    get_all_games,
    get_all_games_for_user,
    delete_game_by_id,
    delete_games_by_round
)
from app.schemas.game import GameCreate, GameRead, GameUpdateResult

router = APIRouter()

# --------------------------------------------------
# ENDPOINT: Upload de Planilha Excel para Jogos (Rodada agora é parâmetro de query)
# --------------------------------------------------
@router.post("/admin/games/upload-excel", response_model=List[GameRead])
async def upload_games_excel(
    current_admin: Annotated[Any, Depends(get_current_active_admin)],
    round_number: Annotated[int, Query(..., ge=1, le=38, description="Número da rodada para os jogos da planilha.")],
    file: UploadFile = File(...),
    db: Session = Depends(get_session)
):
    """
    Faz upload de uma planilha Excel (.xlsx) com jogos para uma rodada específica e os insere no banco de dados.
    A planilha deve ter as colunas: 'mandante', 'visitante', 'data_hora'.
    """
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de arquivo inválido. Por favor, envie um arquivo .xlsx"
        )

    try:
        workbook = openpyxl.load_workbook(file.file)
        sheet = workbook.active
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Erro ao ler a planilha Excel: {e}. Verifique o formato."
        )

    games_to_create = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if not row or all(cell is None for cell in row):
            continue

        try:
            home_team = str(row[0])
            away_team = str(row[1])
            
            game_datetime_excel = row[2]

            game_datetime: datetime

            if isinstance(game_datetime_excel, datetime):
                game_datetime = game_datetime_excel.replace(tzinfo=None)
            elif isinstance(game_datetime_excel, (int, float)):
                from openpyxl.utils.datetime import from_excel
                game_datetime = from_excel(game_datetime_excel).replace(tzinfo=None)
            else:
                dt_str = str(game_datetime_excel).strip()
                try:
                    game_datetime = datetime.fromisoformat(dt_str)
                except ValueError:
                    try:
                        game_datetime = datetime.strptime(dt_str, '%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        raise ValueError(f"Formato de data/hora desconhecido: '{dt_str}'")

            game_create_data = GameCreate(
                round_number=round_number,
                home_team=home_team,
                away_team=away_team,
                game_datetime=game_datetime
            )
            games_to_create.append(game_create_data)
        except ValueError as ve:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Erro de formato de dado na linha {row_index + 2}: {ve}. Verifique 'data_hora' (formato AAAA-MM-DD HH:MM:SS ou ISO)."
            )
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Erro inesperado na linha {row_index + 2}: {e}."
            )

    created_games = []
    for game_data in games_to_create:
        created_games.append(create_game(game_data, db)) # Usar a função CRUD
    
    return created_games

# --------------------------------------------------
# ENDPOINT: Listar Jogos por Rodada
# --------------------------------------------------
@router.get("/games/{round_number}", response_model=List[GameRead])
async def read_games_by_round(
    round_number: int,
    current_user: Annotated[Any, Depends(get_current_user)], # get_current_user vem do core.security
    db: Session = Depends(get_session)
):
    """
    Retorna todos os jogos de uma rodada específica.
    """
    games = get_games_by_round(round_number, db) # Usar a função CRUD
    if not games:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Nenhum jogo encontrado para a rodada {round_number}."
        )
    return games

# --------------------------------------------------
# ENDPOINT: Atualizar Resultado de Jogo (Admin)
# --------------------------------------------------
@router.put("/admin/games/{game_id}/result", response_model=GameRead)
async def update_game_scores(
    game_id: int,
    game_update: GameUpdateResult,
    current_admin: Annotated[Any, Depends(get_current_active_admin)], # get_current_active_admin vem do core.security
    db: Session = Depends(get_session)
):
    """
    Atualiza o placar e status de um jogo específico (apenas para administradores).
    """
    updated_game = update_game_result(game_id, game_update, db) # Usar a função CRUD
    if not updated_game:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Jogo não encontrado.")
    return updated_game

# --------------------------------------------------
# ENDPOINT: Listar Todos os Jogos (Admin)
# --------------------------------------------------
@router.get("/admin/games", response_model=List[GameRead])
async def read_all_games_admin(
    current_admin: Annotated[Any, Depends(get_current_active_admin)], # get_current_active_admin vem do core.security
    db: Session = Depends(get_session)
):
    """
    Retorna a lista de todos os jogos cadastrados (apenas para administradores).
    """
    games = get_all_games(db) # Usar a função CRUD
    return games

# --------------------------------------------------
# NOVOS ENDPOINTS: EXCLUSÃO DE JOGOS (ADMIN)
# (Exigem que o usuário seja um administrador)
# --------------------------------------------------

@router.delete("/admin/games/{game_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_single_game(
    game_id: int,
    current_admin: Annotated[Any, Depends(get_current_active_admin)], # get_current_active_admin vem do core.security
    db: Session = Depends(get_session)
):
    """
    Deleta um único jogo pelo seu ID (apenas para administradores).
    """
    deleted = delete_game_by_id(game_id, db) # Usar a função CRUD
    if not deleted:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Jogo não encontrado.")
    return Response(status_code=status.HTTP_204_NO_CONTENT) # Retorna Response vazio para 204

@router.delete("/admin/rounds/{round_number}", status_code=status.HTTP_200_OK)
async def delete_round_games(
    round_number: int,
    current_admin: Annotated[Any, Depends(get_current_active_admin)], # get_current_active_admin vem do core.security
    db: Session = Depends(get_session)
):
    """
    Deleta todos os jogos de uma rodada específica (apenas para administradores).
    Retorna o número de jogos deletados.
    """
    deleted_count = delete_games_by_round(round_number, db) # Usar a função CRUD
    if deleted_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Nenhum jogo encontrado ou deletado para a rodada {round_number}."
        )
    return {"message": f"{deleted_count} jogo(s) da rodada {round_number} foram deletado(s) com sucesso."}

# --------------------------------------------------
# NOVO ENDPOINT: Gerar Planilha de Resultados para Download (Admin)
# --------------------------------------------------
@router.get("/admin/games/download-results-template/{round_number}", response_class=StreamingResponse)
async def download_results_template(
    round_number: int,
    current_admin: Annotated[Any, Depends(get_current_active_admin)], # get_current_active_admin vem do core.security
    db: Session = Depends(get_session)
):
    """
    Gera e retorna uma planilha Excel com os jogos de uma rodada específica,
    incluindo colunas para preencher os placares (para administradores).
    """
    games = get_games_by_round(round_number, db) # Usar a função CRUD
    if not games:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Nenhum jogo encontrado para a rodada {round_number} para gerar a planilha."
        )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Resultados Rodada {round_number}"

    headers = ["id_jogo", "rodada", "mandante", "visitante", "data_hora", "placar_mandante", "placar_visitante"]
    sheet.append(headers)

    for game in games:
        game_datetime_str = game.game_datetime.isoformat() + 'Z' if game.game_datetime.tzinfo else game.game_datetime.isoformat()
        
        row_data = [
            game.id,
            game.round_number,
            game.home_team,
            game.away_team,
            game_datetime_str,
            None, # Coluna vazia para placar_mandante
            None  # Coluna vazia para placar_visitante
        ]
        sheet.append(row_data)

    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    filename = f"resultados_rodada_{round_number}.xlsx"
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# --------------------------------------------------
# NOVO ENDPOINT: Upload de Planilha de Resultados (Admin)
# --------------------------------------------------
@router.post("/admin/games/upload-results-excel", response_model=List[GameRead])
async def upload_results_excel(
    current_admin: Annotated[Any, Depends(get_current_active_admin)],
    file: UploadFile = File(...),
    db: Session = Depends(get_session)
):
    """
    Faz upload de uma planilha Excel (.xlsx) com resultados de jogos existentes (por ID)
    e atualiza os placares no banco de dados.
    A planilha deve ter as colunas: 'id_jogo', 'rodada', 'mandante', 'visitante',
    'data_hora', 'placar_mandante', 'placar_visitante'.
    """
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de arquivo inválido. Por favor, envie um arquivo .xlsx"
        )

    try:
        workbook = openpyxl.load_workbook(file.file)
        sheet = workbook.active
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Erro ao ler a planilha Excel: {e}. Verifique o formato."
        )

    updated_games = []
    games_to_update_data = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if not row or all(cell is None for cell in row):
            continue

        try:
            game_id = int(row[0]) # id_jogo é a primeira coluna
            home_score = int(row[5]) if row[5] is not None else None # placar_mandante (coluna 5)
            away_score = int(row[6]) if row[6] is not None else None # placar_visitante (coluna 6)
            
            if home_score is None or away_score is None:
                raise ValueError("Placares do mandante e visitante são obrigatórios e devem ser números inteiros.")

            games_to_update_data.append({
                "id": game_id,
                "home_score": home_score,
                "away_score": away_score,
                "status": GameStatus.FINISHED
            })
        except ValueError as ve:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Erro de formato de dado na linha {row_index + 2}: {ve}. Verifique 'id_jogo', 'placar_mandante' e 'placar_visitante' (devem ser números inteiros)."
            )
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Erro inesperado na linha {row_index + 2}: {e}. Verifique a estrutura da planilha."
            )
    
    for game_data in games_to_update_data:
        game_update_result = GameUpdateResult(
            home_score=game_data["home_score"],
            away_score=game_data["away_score"],
            status=game_data["status"]
        )
        updated_game = update_game_result(game_data["id"], game_update_result, db) # Usar a função CRUD
        if not updated_game:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Jogo com ID {game_data['id']} não encontrado para atualização."
            )
        updated_games.append(updated_game)

    return updated_games

# --------------------------------------------------
# NOVO ENDPOINT: Listar Todos os Jogos (para Usuários Comuns)
# --------------------------------------------------
@router.get("/all", response_model=List[GameRead]) # << MUDANÇA: Novo endpoint /all (para usuários)
async def read_all_games_for_user(
    current_user: Annotated[Any, Depends(get_current_user)], # <<< Não exige admin, apenas usuário logado
    db: Session = Depends(get_session)
):
    """
    Retorna a lista de todos os jogos cadastrados (para usuários comuns).
    Filtra jogos que estão agendados, finalizados ou adiados.
    Não retorna jogos cancelados.
    """
    games = get_all_games_for_user(db) # <<< Usar a nova função CRUD
    return games

# --------------------------------------------------
# ENDPOINT: Listar Todos os Jogos (Admin) - Mantido
# --------------------------------------------------
@router.get("/admin/games", response_model=List[GameRead])
async def read_all_games_admin(
    current_admin: Annotated[Any, Depends(get_current_active_admin)],
    db: Session = Depends(get_session)
):
    """
    Retorna a lista de todos os jogos cadastrados (apenas para administradores).
    """
    games = get_all_games(db) # Esta função CRUD já existe e é para admin
    return games